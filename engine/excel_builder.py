# -*- coding: utf-8 -*-
"""
Excel 리포트 빌더
- Sheet 1: 요약 (등급별 종목 리스트 + 총점)
- Sheet 2: 상세 점수표 (9단계 필터 단계별 점수/통과여부)
- Sheet 3: 시스템 설명 (가중치표 + 필터 우선순위표)
"""

import logging
import os
from datetime import datetime
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

from config.settings import WEIGHTS, FILTER_PIPELINE, GRADE_CUTOFFS, MIN_PASS_SCORE, OUTPUT_DIR, EXCEL_FILENAME_PREFIX
from engine.scoring import ScanResult, PIPELINE_ORDER

logger = logging.getLogger("excel_builder")

FONT_NAME = "Arial"

GRADE_COLORS = {
    "S": "FFD700",  # gold
    "A": "92D050",  # green
    "B": "FFEB84",  # yellow
    "C": "F4B084",  # orange
}

HEADER_FILL = PatternFill("solid", start_color="305496", end_color="305496")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=16, color="1F3864")
SUB_FONT = Font(name=FONT_NAME, size=10, italic=True, color="666666")

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def _style_header_row(ws, row_idx: int, n_cols: int):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _autofit(ws, widths: dict):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def build_summary_sheet(wb: Workbook, results: List[ScanResult], scan_date: str):
    ws = wb.active
    ws.title = "요약"

    ws["A1"] = "OMNI KR STOCK SCANNER - 종목발굴 리포트"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:H1")

    ws["A2"] = f"스캔 기준일: {scan_date}   |   대상: KOSPI200 + KOSDAQ100   |   탈락 기준: {MIN_PASS_SCORE}점 미만"
    ws["A2"].font = SUB_FONT
    ws.merge_cells("A2:H2")

    grade_counts = {g: 0 for g, _ in GRADE_CUTOFFS}
    for r in results:
        if r.grade in grade_counts:
            grade_counts[r.grade] += 1
    summary_line = "  ".join([f"{g}등급 {n}건" for g, n in grade_counts.items()])
    ws["A3"] = f"등급 분포: {summary_line}  (총 {len(results)}건 통과)"
    ws["A3"].font = SUB_FONT
    ws.merge_cells("A3:H3")

    header_row = 5
    headers = ["순위", "종목코드", "종목명", "시장", "현재가", "총점", "등급",
               "거래량", "OBV", "VWAP", "MFI", "MACD", "KDJ", "RSI", "ADX", "볼린저"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=i, value=h)
    _style_header_row(ws, header_row, len(headers))

    key_to_col = {key: 8 + idx for idx, key in enumerate(PIPELINE_ORDER)}  # H=8부터

    for rank, r in enumerate(results, start=1):
        row = header_row + rank
        ws.cell(row=row, column=1, value=rank)
        ws.cell(row=row, column=2, value=r.code)
        ws.cell(row=row, column=3, value=r.name)
        ws.cell(row=row, column=4, value="코스피200" if r.market == "KOSPI200" else "코스닥100")
        ws.cell(row=row, column=5, value=r.close).number_format = "#,##0"
        ws.cell(row=row, column=6, value=r.total_score).number_format = "0.0"

        grade_cell = ws.cell(row=row, column=7, value=r.grade)
        grade_cell.font = Font(name=FONT_NAME, bold=True)
        grade_cell.fill = PatternFill("solid", start_color=GRADE_COLORS.get(r.grade, "FFFFFF"))
        grade_cell.alignment = Alignment(horizontal="center")

        step_map = r.step_dict()
        for key, col in key_to_col.items():
            step = step_map[key]
            cell = ws.cell(row=row, column=col, value=round(step.weighted_score, 1))
            cell.number_format = "0.0"
            if step.passed:
                cell.font = Font(name=FONT_NAME, color="1A7A1A")
            else:
                cell.font = Font(name=FONT_NAME, color="999999")

        for c in range(1, len(headers) + 1):
            ws.cell(row=row, column=c).border = THIN_BORDER

    last_row = header_row + len(results)
    if len(results) > 0:
        score_col = "F"
        rng = f"{score_col}{header_row+1}:{score_col}{last_row}"
        ws.conditional_formatting.add(
            rng,
            CellIsRule(operator="greaterThanOrEqual", formula=["85"], fill=PatternFill("solid", start_color="FFD700")),
        )

    widths = {
        "A": 6, "B": 10, "C": 18, "D": 11, "E": 12, "F": 8, "G": 6,
        "H": 9, "I": 8, "J": 9, "K": 7, "L": 8, "M": 7, "N": 7, "O": 7, "P": 8,
    }
    _autofit(ws, widths)
    ws.freeze_panes = f"A{header_row+1}"


def build_detail_sheet(wb: Workbook, results: List[ScanResult]):
    ws = wb.create_sheet("상세점수표")

    ws["A1"] = "단계별 상세 점수 및 판정 근거"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")

    row = 3
    name_map = {f["key"]: f["name"] for f in FILTER_PIPELINE}

    for r in results:
        ws.cell(row=row, column=1, value=f"[{r.grade}] {r.code} {r.name} (총점 {r.total_score})")
        ws.cell(row=row, column=1).font = Font(name=FONT_NAME, bold=True, size=12)
        ws.merge_cells(f"A{row}:F{row}")
        row += 1

        sub_header = ["단계", "지표", "가중치", "통과여부", "단계점수", "상세"]
        for i, h in enumerate(sub_header, start=1):
            ws.cell(row=row, column=i, value=h)
        _style_header_row(ws, row, len(sub_header))
        row += 1

        for step_idx, key in enumerate(PIPELINE_ORDER, start=1):
            step = r.step_dict()[key]
            ws.cell(row=row, column=1, value=step_idx)
            ws.cell(row=row, column=2, value=name_map[key])
            ws.cell(row=row, column=3, value=WEIGHTS[key])
            pass_cell = ws.cell(row=row, column=4, value="통과" if step.passed else "미통과")
            pass_cell.font = Font(name=FONT_NAME, color="1A7A1A" if step.passed else "CC0000", bold=True)
            ws.cell(row=row, column=5, value=round(step.weighted_score, 2)).number_format = "0.00"
            ws.cell(row=row, column=6, value=step.detail)
            for c in range(1, 7):
                ws.cell(row=row, column=c).border = THIN_BORDER
            row += 1

        row += 1  # 종목간 간격

    widths = {"A": 6, "B": 14, "C": 8, "D": 10, "E": 10, "F": 40}
    _autofit(ws, widths)


def build_system_sheet(wb: Workbook):
    ws = wb.create_sheet("시스템설명")

    ws["A1"] = "스코어링 시스템 구조"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")

    ws["A3"] = "■ 지표별 가중치"
    ws["A3"].font = Font(name=FONT_NAME, bold=True, size=12)
    row = 4
    ws.cell(row=row, column=1, value="지표")
    ws.cell(row=row, column=2, value="가중치(%)")
    _style_header_row(ws, row, 2)
    row += 1
    weight_label = {
        "volume": "거래량", "obv": "OBV", "vwap": "VWAP", "kdj": "KDJ",
        "macd": "MACD", "mfi": "MFI", "rsi": "RSI", "adx": "ADX", "bb": "볼린저밴드",
    }
    for key in PIPELINE_ORDER:
        ws.cell(row=row, column=1, value=weight_label[key])
        ws.cell(row=row, column=2, value=WEIGHTS[key])
        for c in (1, 2):
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="■ 필터 우선순위 파이프라인").font = Font(name=FONT_NAME, bold=True, size=12)
    row += 1
    header = ["단계", "필터", "목적", "중요도"]
    for i, h in enumerate(header, start=1):
        ws.cell(row=row, column=i, value=h)
    _style_header_row(ws, row, 4)
    row += 1
    for idx, f in enumerate(FILTER_PIPELINE, start=1):
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=f["name"])
        ws.cell(row=row, column=3, value=f["purpose"])
        ws.cell(row=row, column=4, value="★" * f["stars"])
        for c in range(1, 5):
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="■ 등급 컷오프").font = Font(name=FONT_NAME, bold=True, size=12)
    row += 1
    ws.cell(row=row, column=1, value="등급")
    ws.cell(row=row, column=2, value="최소점수")
    _style_header_row(ws, row, 2)
    row += 1
    for grade, cutoff in GRADE_CUTOFFS:
        ws.cell(row=row, column=1, value=grade).fill = PatternFill("solid", start_color=GRADE_COLORS.get(grade, "FFFFFF"))
        ws.cell(row=row, column=2, value=cutoff)
        for c in (1, 2):
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1
    ws.cell(row=row, column=1, value="탈락")
    ws.cell(row=row, column=2, value=f"< {MIN_PASS_SCORE}")
    row += 1

    widths = {"A": 14, "B": 16, "C": 22, "D": 12}
    _autofit(ws, widths)


def build_excel_report(results: List[ScanResult], scan_date: str = None) -> str:
    if scan_date is None:
        scan_date = datetime.now().strftime("%Y-%m-%d %H:%M KST")

    wb = Workbook()
    build_summary_sheet(wb, results, scan_date)
    build_detail_sheet(wb, results)
    build_system_sheet(wb)

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    fname = f"{EXCEL_FILENAME_PREFIX}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    path = os.path.join(OUTPUT_DIR, fname)
    wb.save(path)
    logger.info(f"엑셀 리포트 저장: {path}")
    return path
